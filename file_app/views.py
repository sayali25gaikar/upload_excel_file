from django.shortcuts import render, redirect
from .forms import UploadFileForm
from openpyxl import load_workbook
from datetime import datetime
from .models import FileUpload, UserData
from django.shortcuts import get_object_or_404
from django.db.models import Count, F
from django.utils.timezone import now
import json
from django.contrib import messages
# Create your views here.

def validate_data_extract(file):
    """
        validation starts here
    """

    
    wb = load_workbook(file)
    sheet = wb.active

    data = []
    errors = []
    sno_set = set()

    correct_columns = ["Sno", "FirstName", "LastName", "Gender", "DateofBirth"]
    actual_columns = []
    for i in range(len(correct_columns)):
        actual_columns.append(sheet.cell(row=1, column=i+1).value)

    if actual_columns != correct_columns:
        return {"errors": ["Invalid column headers or structure"], "data": None}
    data = []
    errors = []
    number_set = set()
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        sno, first_name, last_name, gender, dob = row
        if not isinstance(sno, str) or sno in sno_set:
            errors.append(f"Row {row_idx}: Invalid or duplicate Sno")
        if not isinstance(first_name, str) or not (0 < len(first_name) <= 50):
            errors.append(f"Row {row_idx}: Invalid FirstName")
        if not isinstance(last_name, str) or not (0 < len(last_name) <= 50):
            errors.append(f"Row {row_idx}: Invalid LastName")
        if gender not in ('M', 'F', 'O'):
            errors.append(f"Row {row_idx}: Gender must be 'M', 'F', or 'O'")
        try:
            dob = datetime.strptime(dob, "%Y-%m-%d").date()
            if dob >= datetime.now().date():
                errors.append(f"Row {row_idx}: DateofBirth must be in the past")
        except (TypeError, ValueError):
            errors.append(f"Row {row_idx}: Invalid DateofBirth format, must be YYYY-MM-DD")

        if not errors:
            number_set.add(sno)
            data.append({
                "sno": sno,
                "first_name": first_name,
                "last_name": last_name,
                "gender": gender,
                "date_of_birth": dob,
            })
    
    return {"errors": errors, "data": data}
    

def index(request):
    if request.method == 'GET':
        form = UploadFileForm()
        return render(request, 'upload_file.html', context={'form': form})
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES.get("file")
            if not uploaded_file.name.endswith(".xlsx"):
                messages.error(request, "Only .xlsx files are allowed.")
                return redirect("index")
            res = validate_data_extract(uploaded_file)
            print(res['errors'])
            if res['errors']:
                return render (request,'error_page.html' , {"error": res['errors']})
            
            file_entry = FileUpload.objects.create(row_count=len(res["data"]))
            UserData.objects.bulk_create([
                UserData(file=file_entry, **row_data) for row_data in res["data"]
            ])

            return redirect("view_all_files")

def view_uploaded_files(request):
    files = FileUpload.objects.all()
    return render(request, "view_page.html", {"files": files})


def view_file_data_with_id(request, file_id):
    file_upload = get_object_or_404(FileUpload, id=file_id)
    user_data = file_upload.user_data.all()
    return render(request, "file_detail_view.html", {"file": file_upload, "user_data": user_data})

def dashboard(request):
    data = UserData.objects.annotate(
        age=(now().year - F("date_of_birth__year"))
    ).values("age", "gender").annotate(count=Count("id"))

    data_json = json.dumps(list(data))

    return render(request, "dashboard.html", {"data_json": data_json})